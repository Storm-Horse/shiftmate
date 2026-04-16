from datetime import datetime, timedelta, date as date_type
from io import BytesIO
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ──────────────────────────────────────────────────────────────────

def compute_hours(start_time, end_time, break_minutes: int, direct_hours=None) -> float:
    if direct_hours is not None:
        return round(float(direct_hours), 2)
    fmt = "%H:%M"
    start = datetime.strptime(start_time, fmt)
    end = datetime.strptime(end_time, fmt)
    if end <= start:  # overnight shift
        end += timedelta(days=1)
    total_minutes = int((end - start).total_seconds() / 60) - break_minutes
    return round(max(0, total_minutes) / 60, 2)


def _iso_to_date(iso: str) -> date_type:
    return datetime.strptime(iso, "%Y-%m-%d").date()


def _fmt_display(d: date_type) -> str:
    return d.strftime("%d %b %Y")


def _week_monday(d: date_type) -> date_type:
    """Return the Monday of the week containing d."""
    return d - timedelta(days=d.weekday())


def _week_sunday(monday: date_type) -> date_type:
    return monday + timedelta(days=6)


def _weeks_in_period(period_start: str, period_end: str):
    """Yield (monday, sunday) pairs covering the period."""
    start = _iso_to_date(period_start)
    end = _iso_to_date(period_end)
    monday = _week_monday(start)
    while monday <= end:
        yield monday, _week_sunday(monday)
        monday += timedelta(days=7)


# ── Styles ───────────────────────────────────────────────────────────────────

def _thin_border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _apply(cell, value=None, bold=False, size=11, align_h="left", align_v="center",
           wrap=False, border=False, number_format=None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if border:
        cell.border = _thin_border()
    if number_format:
        cell.number_format = number_format


# ── Week sheet builder ────────────────────────────────────────────────────────

DAYS = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
ROWS_PER_DAY = 5
COL_WIDTHS = {
    "A": 10.43, "B": 8.43, "C": 12.71, "D": 13.14, "E": 49.43,
    "F": 11.43, "G": 10.57, "H": 4.14, "I": 4.43, "J": 3.86,
    "K": 7.43,  "L": 8.0,
}


def _build_week_sheet(ws, user, monday: date_type, sunday: date_type, week_shifts):
    """Populate ws with one week of timesheet data matching the original format."""

    border = _thin_border()

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 29.25
    ws.row_dimensions[2].height = 12.75
    ws.row_dimensions[3].height = 8.25
    ws.row_dimensions[4].height = 15.0
    ws.row_dimensions[5].height = 15.0
    for r in range(6, 42):
        ws.row_dimensions[r].height = 12.75
    ws.row_dimensions[41].height = 18.0

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    _apply(ws["A1"], "TIME SHEET", bold=True, size=16, align_h="center")

    # ── Row 2: Name / Employer / Week ending ──────────────────────────────────
    ws.merge_cells("A2:L2")
    employer = user.employer or ""
    week_ending = sunday.strftime("%d/%m/%Y")
    header_text = (
        f"NAME :   {user.name.upper()}"
        f"{'        ' + employer.upper() if employer else ''}"
        f"        WEEK ENDING :   {week_ending}"
    )
    _apply(ws["A2"], header_text, bold=True, size=12)

    # ── Rows 4–5: Column headers (double-row) ─────────────────────────────────
    # Row 4
    for col, val in [
        ("B", "JOB No"), ("C", "CLIENT"), ("D", "JOB NAME"),
        ("E", "DESCRIPTION OF WORK CARRIED OUT"),
        ("F", "Depart from"), ("G", "Arrive at"),
        ("H", "1.0"), ("I", "1.5"), ("J", "2.0"),
        ("K", "TRAV."), ("L", "TRAV."),
    ]:
        c = ws[f"{col}4"]
        size = 8 if col in ("F", "G") else (10 if col in ("H", "I", "J", "K", "L") else 11)
        align_h = "center" if col in ("H", "I", "J", "K", "L") else "left"
        _apply(c, val, bold=True, size=size, align_h=align_h, border=True)

    # Row 5
    _apply(ws["A5"], "DATE", bold=True, size=12, border=True)
    _apply(ws["F5"], "Depart Time", bold=True, size=8, border=True)
    _apply(ws["G5"], "Arrival Time", bold=True, size=8, border=True)
    _apply(ws["K5"], "30KM", bold=True, size=10, align_h="center", border=True)
    _apply(ws["L5"], "50KM", bold=True, size=10, align_h="center", border=True)

    # ── Day blocks (rows 6–40) ────────────────────────────────────────────────
    # Group shifts by date (iso string)
    by_date = defaultdict(list)
    for shift in week_shifts:
        by_date[shift.date].append(shift)

    for day_idx, day_name in enumerate(DAYS):
        day_date = monday + timedelta(days=day_idx)
        row_start = 6 + day_idx * ROWS_PER_DAY

        # Day name label in col A (first row of block)
        is_weekend = day_idx >= 5
        _apply(
            ws.cell(row=row_start, column=1),
            day_name,
            bold=is_weekend,
            size=8,
            align_h="center",
            border=True,
        )

        # Apply border to all cells in the day block
        for r in range(row_start, row_start + ROWS_PER_DAY):
            for c in range(1, 13):  # A–L
                ws.cell(row=r, column=c).border = border

        # Fill in shifts for this day (up to ROWS_PER_DAY)
        day_shifts = by_date.get(day_date.strftime("%Y-%m-%d"), [])
        for i, shift in enumerate(day_shifts[:ROWS_PER_DAY]):
            r = row_start + i
            hours = compute_hours(
                shift.start_time, shift.end_time,
                shift.break_minutes, shift.direct_hours,
            )

            if shift.job_name:
                _apply(ws.cell(row=r, column=4), shift.job_name, size=10, align_h="center")
            if shift.notes:
                _apply(ws.cell(row=r, column=5), shift.notes, size=10)
            if shift.start_time:
                _apply(ws.cell(row=r, column=6), shift.start_time, size=10, align_h="center")
            if shift.end_time:
                _apply(ws.cell(row=r, column=7), shift.end_time, size=10, align_h="center")

            h_cell = ws.cell(row=r, column=8)
            h_cell.value = hours
            h_cell.font = Font(size=10)
            h_cell.alignment = Alignment(horizontal="center", vertical="center")
            h_cell.number_format = "0.##"  # shows 7.5 not 7.50, matches original

    # ── Row 41: NOTES ─────────────────────────────────────────────────────────
    ws.merge_cells("A41:L41")
    _apply(ws["A41"], "NOTES", bold=True, size=12)

    # ── Row 42: Total hours summary ───────────────────────────────────────────
    ws.row_dimensions[42].height = 15.0
    total = sum(
        compute_hours(s.start_time, s.end_time, s.break_minutes, s.direct_hours)
        for s in week_shifts
    )
    ws.merge_cells("A42:G42")
    _apply(ws["A42"], f"TOTAL HOURS :   {round(total, 2)}", bold=True, size=11, border=True)

    # Job breakdown
    job_totals = defaultdict(float)
    for s in week_shifts:
        job_totals[s.job_name] += compute_hours(
            s.start_time, s.end_time, s.break_minutes, s.direct_hours
        )
    ws.row_dimensions[43].height = 12.75
    summary = "   ".join(f"{job}: {round(h, 2)}h" for job, h in sorted(job_totals.items()))
    ws.merge_cells("A43:L43")
    _apply(ws["A43"], summary, size=10)


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(user, shifts, period_start: str, period_end: str) -> bytes:
    """Generate an xlsx matching the original timesheet format.
    One sheet per Mon–Sun week covering the period.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet; we'll add our own

    # Index shifts by date string for fast lookup
    by_date = defaultdict(list)
    for shift in shifts:
        by_date[shift.date].append(shift)

    weeks = list(_weeks_in_period(period_start, period_end))
    for monday, sunday in weeks:
        # Collect shifts that fall within this Mon–Sun week
        week_shifts = [
            s for s in shifts
            if monday <= _iso_to_date(s.date) <= sunday
        ]

        sheet_name = f"Week {monday.strftime('%-d %b')}"
        ws = wb.create_sheet(title=sheet_name)
        _build_week_sheet(ws, user, monday, sunday, week_shifts)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
